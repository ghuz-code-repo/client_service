# app/routes.py
import time
import math
import json
import io
import os
import uuid
import tempfile
from datetime import datetime
from threading import Thread
from openpyxl import load_workbook, Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from flask import (render_template, request, Blueprint, abort, flash, redirect,
                   url_for, jsonify, current_app, send_file, g)
from sqlalchemy.orm import selectinload, joinedload
from sqlalchemy import text
from flask_mail import Message
from werkzeug.utils import secure_filename
from .extensions import db
from .models import (User, EstateDealsContacts, EstateDeals, EstateSells, Client,
                     Application, Defect, ApplicationLog, ResponsiblePerson, EstateHouses, responsible_assignments,
                     DefectType, EmailLog, ApplicationType)
from .email_utils import generate_and_send_email
from .decorators import permission_required, admin_required, auth_required  # Import auth_required
from sqlalchemy import or_

main = Blueprint('main', __name__)


@main.route('/health')
def health():
    """Health check endpoint for Docker health checks."""
    return {'status': 'ok', 'service': 'client-service'}, 200


@main.route('/api/sync/permissions')
def sync_permissions():
    """
    Endpoint для синхронизации permissions с auth-service (PULL-модель).
    Auth-service вызывает этот endpoint для получения списка permissions.
    """
    permissions = [
        # Applications
        {"name": "client-service.applications.view.all", "displayName": "Просмотр всех заявок", 
         "description": "Разрешение на просмотр всех заявок в системе", "category": "applications"},
        {"name": "client-service.applications.view.own", "displayName": "Просмотр своих заявок", 
         "description": "Разрешение на просмотр только созданных пользователем заявок", "category": "applications"},
        {"name": "client-service.applications.view.responsible", "displayName": "Просмотр заявок где ответственный", 
         "description": "Разрешение на просмотр заявок где пользователь назначен ответственным", "category": "applications"},
        {"name": "client-service.applications.create", "displayName": "Создание заявок",
         "description": "Разрешение на создание новых заявок", "category": "applications"},
        {"name": "client-service.applications.edit", "displayName": "Редактирование заявок",
         "description": "Разрешение на редактирование заявок", "category": "applications"},
        {"name": "client-service.applications.delete", "displayName": "Удаление заявок",
         "description": "Разрешение на удаление заявок", "category": "applications"},
        {"name": "client-service.applications.assign", "displayName": "Назначение ответственных",
         "description": "Разрешение на назначение ответственных", "category": "applications"},
        {"name": "client-service.applications.status.change", "displayName": "Изменение статуса",
         "description": "Разрешение на изменение статуса заявок", "category": "applications"},
        {"name": "client-service.applications.export", "displayName": "Экспорт заявок",
         "description": "Разрешение на экспорт заявок в Excel", "category": "applications"},
        {"name": "client-service.applications.import", "displayName": "Импорт заявок",
         "description": "Разрешение на массовый импорт и обновление заявок из Excel", "category": "applications"},
        
        # Responsible
        {"name": "client-service.responsible.view", "displayName": "Просмотр ответственных",
         "description": "Разрешение на просмотр ответственных лиц", "category": "responsible"},
        {"name": "client-service.responsible.create", "displayName": "Создание ответственных",
         "description": "Разрешение на создание ответственных лиц", "category": "responsible"},
        {"name": "client-service.responsible.edit", "displayName": "Редактирование ответственных",
         "description": "Разрешение на редактирование ответственных лиц", "category": "responsible"},
        {"name": "client-service.responsible.delete", "displayName": "Удаление ответственных",
         "description": "Разрешение на удаление ответственных лиц", "category": "responsible"},
        
        # Admin
        {"name": "client-service.admin.panel", "displayName": "Панель администратора",
         "description": "Доступ к панели администратора", "category": "admin"},
        {"name": "client-service.admin.users", "displayName": "Управление пользователями",
         "description": "Управление пользователями системы", "category": "admin"},
        {"name": "client-service.admin.settings", "displayName": "Настройки системы",
         "description": "Управление настройками системы", "category": "admin"},
        {"name": "client-service.admin.logs", "displayName": "Просмотр логов",
         "description": "Доступ к системным логам", "category": "admin"},
        
        # Reports
        {"name": "client-service.reports.view", "displayName": "Просмотр отчетов",
         "description": "Разрешение на просмотр и генерацию отчетов", "category": "reports"},
        {"name": "client-service.reports.download", "displayName": "Скачивание отчетов",
         "description": "Разрешение на скачивание отчетов в Excel", "category": "reports"},
    ]
    
    return {
        'success': True,
        'permissions': permissions,
        'service_key': 'client-service',
        'total': len(permissions)
    }, 200


def send_email_async(app, application_id):
    """
    Функция-обертка для запуска отправки email в отдельном потоке
    с собственным контекстом приложения.
    """
    with app.app_context():
        generate_and_send_email(application_id)


class SQLPagination:
    """Простой класс для имитации объекта пагинации Flask-SQLAlchemy."""

    def __init__(self, items, page, per_page, total):
        self.items = items
        self.page = page
        self.per_page = per_page
        self.total = total

    @property
    def pages(self):
        return math.ceil(self.total / self.per_page) if self.per_page > 0 else 0

    @property
    def has_prev(self):
        return self.page > 1

    @property
    def has_next(self):
        return self.page < self.pages

    @property
    def prev_num(self):
        return self.page - 1

    @property
    def next_num(self):
        return self.page + 1

    def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
        last = 0
        for num in range(1, self.pages + 1):
            if num <= left_edge or \
                    (self.page - left_current - 1 < num < self.page + right_current) or \
                    num > self.pages - right_edge:
                if last + 1 != num:
                    yield None
                yield num
                last = num


def build_client_filters(args):
    """
    Парсит query-параметры и строит SQL-фильтры для страницы клиентов.
    
    Возвращает (where_parts: list[str], params: dict, has_app_filter: bool)
    where_parts — дополнительные условия для WHERE
    params — параметры для SQL
    has_app_filter — нужен ли EXISTS подзапрос к applications
    """
    where_parts = []
    app_conditions = []
    params = {}

    # --- Фильтры по клиенту/договору ---
    search_query = args.get('search', '').strip()
    if search_query:
        where_parts.append("(c.contacts_buy_name LIKE :search OR c.contacts_buy_phones LIKE :search OR d.agreement_number LIKE :search)")
        params['search'] = f'%{search_query}%'

    client_id = args.get('client_id', '').strip()
    if client_id:
        where_parts.append("c.id = :client_id")
        params['client_id'] = client_id

    complex_name = args.get('complex_name', '').strip()
    if complex_name:
        where_parts.append("h.complex_name = :complex_name")
        params['complex_name'] = complex_name

    house_name = args.get('house_name', '').strip()
    if house_name:
        where_parts.append("h.name = :house_name")
        params['house_name'] = house_name

    # --- Фильтры по заявкам (через EXISTS подзапрос) ---
    app_status = args.get('app_status', '').strip()
    if app_status:
        app_conditions.append("a.status = :app_status")
        params['app_status'] = app_status

    app_type = args.get('app_type', '').strip()
    if app_type:
        app_conditions.append("a.application_type = :app_type")
        params['app_type'] = app_type

    responsible_id = args.get('responsible_id', '').strip()
    if responsible_id:
        app_conditions.append("a.responsible_person_id = :responsible_id")
        params['responsible_id'] = responsible_id

    date_from = args.get('date_from', '').strip()
    if date_from:
        app_conditions.append("a.created_at >= :date_from")
        params['date_from'] = date_from

    date_to = args.get('date_to', '').strip()
    if date_to:
        app_conditions.append("a.created_at < date(:date_to, '+1 day')")
        params['date_to'] = date_to

    overdue = args.get('overdue', '').strip()
    if overdue == 'yes':
        app_conditions.append("a.due_date IS NOT NULL AND a.completed_at IS NULL AND a.due_date < datetime('now')")
    elif overdue == 'no':
        app_conditions.append("(a.due_date IS NULL OR a.completed_at IS NOT NULL OR a.due_date >= datetime('now'))")

    app_source = args.get('app_source', '').strip()
    if app_source:
        app_conditions.append("a.source = :app_source")
        params['app_source'] = app_source

    has_app_filter = len(app_conditions) > 0
    if has_app_filter:
        exists_clause = "EXISTS (SELECT 1 FROM applications a WHERE a.client_id = c.id AND " + " AND ".join(app_conditions) + ")"
        where_parts.append(exists_clause)

    return where_parts, params, has_app_filter


@main.route('/')
@auth_required(any_of=['client-service.applications.view.all', 'client-service.applications.view.own', 'client-service.applications.view.responsible'])
def index():
    page = request.args.get('page', 1, type=int)
    per_page = 100
    offset = (page - 1) * per_page

    # --- Построение фильтров ---
    extra_where, params, has_app_filter = build_client_filters(request.args)

    # Базовые условия (клиент с именем, телефоном и договором)
    base_conditions = [
        "c.contacts_buy_name IS NOT NULL", "c.contacts_buy_name!=''",
        "c.contacts_buy_phones IS NOT NULL", "c.contacts_buy_phones!=''",
        "d.agreement_number IS NOT NULL", "d.agreement_number!=''"
    ]

    # Если фильтруем по ЖК/дому, нужен JOIN на sells/houses уже в подзапросе пагинации
    need_house_join = bool(request.args.get('complex_name', '').strip() or request.args.get('house_name', '').strip())

    from_clause = "FROM estate_deals_contacts c JOIN estate_deals d ON c.id=d.contacts_buy_id"
    if need_house_join:
        from_clause += " LEFT JOIN estate_sells s ON d.estate_sell_id=s.estate_sell_id LEFT JOIN estate_houses h ON s.house_id=h.house_id"

    all_conditions = base_conditions + extra_where
    where_clause = "WHERE " + " AND ".join(all_conditions)

    # COUNT
    count_sql = f"SELECT COUNT(DISTINCT c.id) {from_clause} {where_clause}"
    total_clients = db.session.execute(text(count_sql), params).scalar() or 0

    # DATA
    data_sql = f"""
        SELECT c.id AS client_id,c.contacts_buy_name,c.contacts_buy_phones,
               d.agreement_number,d.deal_sum,d.finances_income_reserved,
               s.estate_floor,s.estate_riser,s.geo_flatnum,s.estate_rooms,
               h.complex_name,h.name as house_name
        FROM estate_deals_contacts c
        JOIN estate_deals d ON c.id=d.contacts_buy_id
        LEFT JOIN estate_sells s ON d.estate_sell_id=s.estate_sell_id
        LEFT JOIN estate_houses h ON s.house_id=h.house_id
        JOIN(SELECT DISTINCT c.id {from_clause} {where_clause} ORDER BY c.contacts_buy_name LIMIT :limit OFFSET :offset) AS page_ids ON c.id=page_ids.id
        WHERE d.agreement_number IS NOT NULL AND d.agreement_number!=''
    """
    params['limit'], params['offset'] = per_page, offset
    all_data = db.session.execute(text(data_sql), params).mappings().all()
    clients_data, ordered_client_ids = {}, []
    for row in all_data:
        client_id = row['client_id']
        if client_id not in clients_data:
            ordered_client_ids.append(client_id)
            contact_info = {'id': row.get('client_id'), 'contacts_buy_name': row.get('contacts_buy_name'),
                            'contacts_buy_phones': row.get('contacts_buy_phones')}
            clients_data[client_id] = {'contact': contact_info, 'deals': []}
        clients_data[client_id]['deals'].append(row)

    client_list = []
    for cid in ordered_client_ids:
        data = clients_data.get(cid)
        if data:
            structured_deals = []
            for deal_row in data['deals']:
                sell_obj = type('obj', (object,), {'estate_floor': deal_row.get('estate_floor'),
                                                   'estate_riser': deal_row.get('estate_riser'),
                                                   'geo_flatnum': deal_row.get('geo_flatnum'),
                                                   'estate_rooms': deal_row.get('estate_rooms'),
                                                   'house': type('obj', (object,),
                                                                 {'complex_name': deal_row.get('complex_name'),
                                                                  'name': deal_row.get('house_name')})})
                structured_deals.append(type('obj', (object,), {'agreement_number': deal_row.get('agreement_number'),
                                                                'deal_sum': deal_row.get('deal_sum'),
                                                                'finances_income_reserved': deal_row.get(
                                                                    'finances_income_reserved'), 'sell': sell_obj}))
            contact_obj = type('obj', (object,), data['contact'])
            client_list.append(Client(contact_obj, structured_deals))

    pagination = SQLPagination(client_list, page, per_page, total_clients)
    
    # Получаем данные для модального окна и фильтров
    application_types = ApplicationType.query.order_by(ApplicationType.name).all()
    defect_types_query = DefectType.query.order_by(DefectType.name).all()
    defect_types = [{'id': dt.id, 'name': dt.name} for dt in defect_types_query]
    responsible_persons = ResponsiblePerson.query.order_by(ResponsiblePerson.full_name).all()

    # Список статусов для фильтра
    app_statuses = ['В работе', 'Выполнено', 'Частично выполнено', 'Закрыто', 'Отклонено']
    # Список источников для фильтра
    app_sources = ['Звонок', 'Email', 'Личный визит', 'Сайт', 'Другое']

    # Собираем активные фильтры для передачи в пагинацию
    filter_keys = ['search', 'client_id', 'complex_name', 'house_name',
                   'app_status', 'app_type', 'responsible_id', 'date_from', 'date_to',
                   'overdue', 'app_source']
    filter_params = {k: request.args.get(k, '') for k in filter_keys if request.args.get(k, '').strip()}

    return render_template('index.html', 
                         clients=client_list, 
                         pagination=pagination, 
                         search_query=request.args.get('search', ''),
                         filter_params=filter_params,
                         application_types=application_types,
                         app_statuses=app_statuses,
                         app_sources=app_sources,
                         defect_types=defect_types,
                         responsible_persons=responsible_persons)


@main.route('/client/<signed_int:client_id>')
@auth_required(any_of=['client-service.applications.view.all', 'client-service.applications.view.own', 'client-service.applications.view.responsible'])
def client_card(client_id):
    contact = EstateDealsContacts.query.get_or_404(client_id)
    deals_for_client = contact.deals.options(selectinload(EstateDeals.sell).selectinload(EstateSells.house)).all()
    applications = Application.query.filter_by(client_id=client_id).order_by(Application.created_at.desc()).all()
    client = Client(contact, deals_for_client)

    defect_types_query = DefectType.query.order_by(DefectType.name).all()
    defect_types = [{'id': dt.id, 'name': dt.name} for dt in defect_types_query]

    application_types = ApplicationType.query.order_by(ApplicationType.name).all()
    application_statuses = current_app.config['APPLICATION_STATUSES']

    warranty_info = {}
    for deal in deals_for_client:
        if deal.sell and deal.sell.house:
            house_key = deal.sell.house.name
            if house_key and house_key not in warranty_info:
                warranty_info[house_key] = {
                    'house_date': deal.sell.house.warranty_house_end_date,
                    'apartments_date': deal.sell.house.warranty_apartments_end_date
                }

    current_date = datetime.now().date()

    return render_template('client_card.html',
                           client=client,
                           applications=applications,
                           defect_types=defect_types,
                           application_types=application_types,
                           application_statuses=application_statuses,
                           warranty_info=warranty_info,
                           current_date=current_date,
                           client_comment=contact.client_comment)


@main.route('/application/<int:app_id>')
@auth_required(any_of=['client-service.applications.view.all', 'client-service.applications.view.own', 'client-service.applications.view.responsible'])
def application_card(app_id):
    """Карточка заявки — детальный просмотр"""
    from .auth_utils import has_permission, get_current_user_id, get_or_create_local_user

    app_obj = Application.query.options(
        joinedload(Application.client),
        joinedload(Application.responsible_person),
        joinedload(Application.creator),
    ).get_or_404(app_id)

    # Проверка доступа: все / свои / ответственный / NC-заявки
    if not has_permission('client-service.applications.view.all'):
        current_gateway_user_id = get_current_user_id()
        local_user = get_or_create_local_user(commit=True) if current_gateway_user_id else None
        allowed = False
        if local_user and has_permission('client-service.applications.view.own') and app_obj.creator_id == local_user.id:
            allowed = True
        if current_gateway_user_id and has_permission('client-service.applications.view.responsible'):
            rp = ResponsiblePerson.query.filter_by(gateway_user_id=current_gateway_user_id).first()
            if rp and rp.id == app_obj.responsible_person_id:
                allowed = True
        # NC/SYSTEM-001 заявки доступны всем авторизованным
        if app_obj.agreement_number and (app_obj.agreement_number.startswith('NC-') or app_obj.agreement_number == 'SYSTEM-001'):
            allowed = True
        if not allowed:
            abort(403)

    defects = Defect.query.filter_by(application_id=app_id).all()
    logs = ApplicationLog.query.filter_by(application_id=app_id).options(
        joinedload(ApplicationLog.author)
    ).order_by(ApplicationLog.timestamp.desc()).all()

    application_statuses = current_app.config['APPLICATION_STATUSES']

    return render_template('application_card.html',
                           app=app_obj,
                           defects=defects,
                           logs=logs,
                           application_statuses=application_statuses)


def _apply_app_filters(query, args):
    """Применяет все фильтры заявок к запросу. Единая точка для листинга и экспорта.
    Возвращает (query, applied_filters_dict) — dict с непустыми значениями фильтров."""
    filters = {}

    # Фильтр по статусу
    status = args.get('status', '')
    if status:
        query = query.filter(Application.status == status)
        filters['status'] = status

    # Фильтр по типу заявки
    app_type = args.get('type', '')
    if app_type:
        query = query.filter(Application.application_type == app_type)
        filters['type'] = app_type

    # Фильтр по ответственному
    responsible_id = args.get('responsible_id', '')
    if responsible_id:
        query = query.filter(Application.responsible_person_id == responsible_id)
        filters['responsible_id'] = responsible_id

    # Фильтр по источнику
    source = args.get('source', '')
    if source:
        query = query.filter(Application.source == source)
        filters['source'] = source

    # Фильтр по ЖК
    housing_complex = args.get('housing_complex', '')
    if housing_complex:
        query = query.filter(Application.housing_complex == housing_complex)
        filters['housing_complex'] = housing_complex

    # Фильтр по дому
    house_number = args.get('house_number', '')
    if house_number:
        query = query.filter(Application.house_number == house_number)
        filters['house_number'] = house_number

    # Фильтр по ID заявки
    app_id_filter = args.get('app_id', '').strip()
    if app_id_filter:
        try:
            query = query.filter(Application.id == int(app_id_filter))
            filters['app_id'] = app_id_filter
        except ValueError:
            pass

    # Фильтр по датам создания
    date_from = args.get('date_from', '')
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Application.created_at >= date_from_obj)
            filters['date_from'] = date_from
        except ValueError:
            pass

    date_to = args.get('date_to', '')
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Application.created_at <= date_to_obj)
            filters['date_to'] = date_to
        except ValueError:
            pass

    # Поиск по клиенту, телефону или номеру договора
    search = args.get('search', '')
    if search:
        query = query.join(Application.client).filter(
            or_(
                EstateDealsContacts.contacts_buy_name.contains(search),
                EstateDealsContacts.contacts_buy_phones.contains(search),
                Application.agreement_number.contains(search)
            )
        )
        filters['search'] = search

    # Фильтр по просроченным
    overdue = args.get('overdue', '')
    if overdue == 'yes':
        query = query.filter(
            Application.due_date.isnot(None),
            Application.due_date < datetime.now(),
            Application.completed_at.is_(None)
        )
        filters['overdue'] = 'yes'
    elif overdue == 'no':
        query = query.filter(
            or_(
                Application.due_date.is_(None),
                Application.due_date >= datetime.now(),
                Application.completed_at.isnot(None)
            )
        )
        filters['overdue'] = 'no'

    # Сортировка
    sort = args.get('sort', 'created_desc')
    if sort == 'created_desc':
        query = query.order_by(Application.created_at.desc())
    elif sort == 'created_asc':
        query = query.order_by(Application.created_at.asc())
    elif sort == 'due_desc':
        query = query.order_by(Application.due_date.desc().nullslast())
    elif sort == 'due_asc':
        query = query.order_by(Application.due_date.asc().nullsfirst())
    elif sort == 'id_desc':
        query = query.order_by(Application.id.desc())
    elif sort == 'id_asc':
        query = query.order_by(Application.id.asc())
    else:
        query = query.order_by(Application.created_at.desc())
    filters['sort'] = sort

    return query, filters


@main.route('/export-applications')
@auth_required(permission='client-service.applications.export')
def export_applications():
    """Экспорт заявок в Excel с учетом всех фильтров"""
    # Базовый запрос с предзагрузкой связанных данных
    query = Application.query.options(
        joinedload(Application.client),
        joinedload(Application.responsible_person),
        joinedload(Application.creator),
        joinedload(Application.defects)
    )

    # Фильтрация заявок в зависимости от разрешений
    from .auth_utils import has_permission, get_current_user_id, get_or_create_local_user
    
    # Если нет разрешения на просмотр ВСЕХ заявок, фильтруем по другим разрешениям
    if not has_permission('client-service.applications.view.all'):
        current_gateway_user_id = get_current_user_id()
        local_user = get_or_create_local_user(commit=True) if current_gateway_user_id else None
        
        if local_user:
            filter_conditions = []
            
            if has_permission('client-service.applications.view.own'):
                filter_conditions.append(Application.creator_id == local_user.id)
            
            if has_permission('client-service.applications.view.responsible'):
                responsible_person = ResponsiblePerson.query.filter_by(gateway_user_id=current_gateway_user_id).first()
                if responsible_person:
                    filter_conditions.append(Application.responsible_person_id == responsible_person.id)
            
            if filter_conditions:
                query = query.filter(or_(*filter_conditions))
            else:
                query = query.filter(Application.id == -1)
        else:
            query = query.filter(Application.id == -1)
    
    # Применяем все фильтры через общую функцию
    query, filters = _apply_app_filters(query, request.args)

    # Получаем все заявки без пагинации для экспорта
    apps = query.all()
    
    if not apps:
        flash('Не найдено заявок для экспорта с указанными фильтрами.', 'info')
        return redirect(url_for('main.applications'))

    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    # Создаем заголовок с информацией о фильтрах
    filter_info = []
    if filters.get('status'):
        filter_info.append(f"Статус: {filters['status']}")
    if filters.get('type'):
        filter_info.append(f"Тип: {filters['type']}")
    if filters.get('responsible_id'):
        rp = ResponsiblePerson.query.get(filters['responsible_id'])
        filter_info.append(f"Ответственный: {rp.full_name if rp else filters['responsible_id']}")
    if filters.get('source'):
        filter_info.append(f"Источник: {filters['source']}")
    if filters.get('housing_complex'):
        filter_info.append(f"ЖК: {filters['housing_complex']}")
    if filters.get('house_number'):
        filter_info.append(f"Дом: {filters['house_number']}")
    if filters.get('app_id'):
        filter_info.append(f"ID заявки: {filters['app_id']}")
    if filters.get('date_from'):
        filter_info.append(f"С даты: {filters['date_from']}")
    if filters.get('date_to'):
        filter_info.append(f"По дату: {filters['date_to']}")
    if filters.get('search'):
        filter_info.append(f"Поиск: {filters['search']}")
    if filters.get('overdue') == 'yes':
        filter_info.append("Только просроченные")
    elif filters.get('overdue') == 'no':
        filter_info.append("Не просроченные")
    
    if filter_info:
        ws.append([f"Отчет по заявкам. Фильтры: {', '.join(filter_info)}"])
        ws.merge_cells('A1:N1')
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.append([])  # Пустая строка
    
    # Заголовки колонок
    headers = ["ID Заявки", "Дата создания", "Статус", "Тип заявки", "№ Договора", 
               "ФИО Клиента", "Телефон клиента", "ЖК", "Дом", "Подъезд", "Номер квартиры",
               "Ответственный", "Email ответственного", "Создатель заявки", "Источник", 
               "Срок выполнения", "Дата завершения", "Просрочено", "Дата последнего изменения", 
               "Последний комментарий", "Комментарий к заявке", "Дефекты (Тип: Комментарий)"]
    ws.append(headers)
    
    # Стилизация заголовков
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Заполнение данными
    for app in apps:
        # Получаем данные о недвижимости
        complex_name = "N/A"
        house_name = "N/A"
        entrance = "N/A"
        flat_num = "N/A"
        
        if app.client and app.agreement_number:
            deal = EstateDeals.query.filter_by(
                agreement_number=app.agreement_number,
                contacts_buy_id=app.client.id
            ).first()
            
            if deal and deal.sell:
                sell = deal.sell
                entrance = sell.geo_house_entrance if sell.geo_house_entrance else "N/A"
                flat_num = sell.geo_flatnum if sell.geo_flatnum else "N/A"
                
                if sell.house:
                    complex_name = sell.house.complex_name if sell.house.complex_name else "N/A"
                    house_name = sell.house.name if sell.house.name else "N/A"
        
        defects_str = "; ".join([f"{d.defect_type}: {d.description}" for d in app.defects])
        is_overdue = "Да" if app.is_overdue else "Нет"
        
        # Получаем последний комментарий из логов
        last_log = ApplicationLog.query.filter_by(application_id=app.id).order_by(ApplicationLog.timestamp.desc()).first()
        last_comment = last_log.comment if last_log else "N/A"
        
        row_data = [
            app.id, 
            app.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            app.status, 
            app.application_type,
            app.agreement_number, 
            app.client.contacts_buy_name if app.client else "N/A",
            app.client.contacts_buy_phones if app.client else "N/A",
            complex_name, 
            house_name, 
            entrance, 
            flat_num,
            app.responsible_person.full_name if app.responsible_person else "Не назначен",
            app.responsible_person.email if app.responsible_person else "N/A",
            app.creator.username if app.creator else "Система",
            app.source if app.source else "Не указан",
            app.due_date.strftime('%Y-%m-%d') if app.due_date else "N/A",
            app.completed_at.strftime('%Y-%m-%d %H:%M:%S') if app.completed_at else "N/A",
            is_overdue,
            app.last_status_change.strftime('%Y-%m-%d %H:%M:%S') if app.last_status_change else "N/A",
            last_comment,
            app.comment, 
            defects_str
        ]
        ws.append(row_data)

    # Автоматическая настройка ширины колонок
    for col in ws.columns:
        max_length = 0
        # Получаем букву колонки из строки заголовков (3-я строка, индекс 2),
        # чтобы избежать ошибки с объединенной ячейкой в первой строке.
        column = col[2].column_letter
        for cell in col:
            try:
                if cell.value:  # Проверяем, что ячейка не пустая
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Сохранение в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Формирование имени файла
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"applications_export_{timestamp}.xlsx"
    
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==================== ИМПОРТ ЗАЯВОК ====================

# Обновляемые колонки (заголовок Excel → внутреннее имя)
IMPORT_UPDATABLE = {
    'Статус': 'status',
    'Ответственный': 'responsible',
    'Источник': 'source',
    'Срок выполнения': 'due_date',
    'Последний комментарий': 'log_comment',
    'Комментарий к заявке': 'comment',
}

VALID_STATUSES = {'В работе', 'Выполнено', 'Частично выполнено', 'Закрыто', 'Отклонено'}
VALID_STATUSES_LIST = ['В работе', 'Выполнено', 'Частично выполнено', 'Закрыто', 'Отклонено']
MAX_IMPORT_FILE_SIZE = 20 * 1024 * 1024  # 20 МБ

# Колонки шаблона импорта (совпадают с экспортом, чтобы файл экспорта
# можно было использовать напрямую как файл импорта)
IMPORT_TEMPLATE_HEADERS = [
    "ID Заявки", "Дата создания", "Статус", "Тип заявки", "№ Договора",
    "ФИО Клиента", "Телефон клиента", "ЖК", "Дом", "Подъезд", "Номер квартиры",
    "Ответственный", "Email ответственного", "Создатель заявки", "Источник",
    "Срок выполнения", "Дата завершения", "Просрочено", "Дата последнего изменения",
    "Последний комментарий", "Комментарий к заявке", "Дефекты (Тип: Комментарий)"
]

# Индексы (1-based) колонок, которые влияют на импорт (редактируемые)
IMPORT_EDITABLE_COLS = {1, 3, 4, 6, 7, 8, 9, 12, 15, 16, 20, 21}
# Остальные колонки read-only: 2, 5, 10, 11, 13, 14, 17, 18, 19, 22


def _create_nc_contact_and_deal(contact_name, contact_phone, client_comment=None):
    """Создаёт NC-контакт и NC-договор. Возвращает (client, agreement_number).
    Вызывается внутри транзакции — НЕ делает commit."""
    min_contact_id = db.session.query(db.func.min(EstateDealsContacts.id)).scalar() or 0
    new_client_id = min(min_contact_id, 0) - 1

    new_client = EstateDealsContacts(
        id=new_client_id,
        contacts_buy_name=contact_name,
        contacts_buy_phones=contact_phone,
        client_comment=client_comment
    )
    db.session.add(new_client)
    db.session.flush()

    agreement_number = f"NC-{abs(new_client.id)}"
    min_deal_id = db.session.query(db.func.min(EstateDeals.id)).scalar() or 0
    new_deal_id = min(min_deal_id, 0) - 1

    new_deal = EstateDeals(
        id=new_deal_id,
        contacts_buy_id=new_client.id,
        agreement_number=agreement_number,
        deal_status_name="Без договора",
        agreement_date=datetime.now().date(),
        deal_sum=0.0,
        finances_income_reserved=0.0
    )
    db.session.add(new_deal)
    db.session.flush()
    return new_client, agreement_number


@main.route('/applications/import-template', methods=['GET'])
@auth_required(permission='client-service.applications.import')
def import_template():
    """Скачивание пустого Excel-шаблона импорта с dropdown-списками.
    Структура колонок идентична экспортному файлу (22 колонки)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Импорт заявок'

    # Лист справочника для источников данных dropdowns
    ws_ref = wb.create_sheet('Справочник')
    # Лист инструкции
    ws_instr = wb.create_sheet('Инструкция')

    # --- Собираем справочные данные ---
    rp_names = sorted([rp.full_name for rp in ResponsiblePerson.query.all()])
    app_type_names = sorted([at.name for at in ApplicationType.query.order_by(ApplicationType.name).all()])
    sources_list = ['Звонок', 'Email', 'Личный визит', 'Сайт', 'Другое']
    complexes = [row[0] for row in db.session.execute(
        db.text('SELECT DISTINCT complex_name FROM estate_houses WHERE complex_name IS NOT NULL AND complex_name != "" ORDER BY complex_name')
    ).fetchall()]
    houses = [row[0] for row in db.session.execute(
        db.text('SELECT DISTINCT name FROM estate_houses WHERE name IS NOT NULL AND name != "" ORDER BY name')
    ).fetchall()]

    # --- Заполняем лист справочника ---
    ref_columns = {
        'A': ('Статусы', VALID_STATUSES_LIST),
        'B': ('Ответственные', rp_names),
        'C': ('Типы заявок', app_type_names),
        'D': ('Источники', sources_list),
        'E': ('ЖК', complexes),
        'F': ('Дома', houses),
    }
    for col_letter, (header, values) in ref_columns.items():
        ws_ref[f'{col_letter}1'] = header
        ws_ref[f'{col_letter}1'].font = Font(bold=True)
        for i, val in enumerate(values, start=2):
            ws_ref[f'{col_letter}{i}'] = val

    # --- Заголовки основного листа (22 колонки = как экспорт) ---
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    gray_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    header_font = Font(bold=True)
    for col_idx, header in enumerate(IMPORT_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = white_fill if col_idx in IMPORT_EDITABLE_COLS else gray_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ширина колонок (22 колонки)
    col_widths = [12, 18, 22, 20, 16, 25, 18, 20, 15, 12, 16, 25, 25, 20, 16, 16, 18, 14, 20, 35, 35, 35]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Data Validation (dropdown-списки) — данные начинаются со строки 2 ---
    max_data_rows = 1000

    def add_ref_validation(ws_target, col_letter_target, col_letter_ref, count, col_idx):
        if count < 1:
            return
        formula = f'Справочник!${col_letter_ref}$2:${col_letter_ref}${count + 1}'
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)
        dv.error = 'Выберите значение из списка'
        dv.errorTitle = 'Недопустимое значение'
        dv.prompt = 'Выберите из списка'
        dv.promptTitle = IMPORT_TEMPLATE_HEADERS[col_idx - 1]
        cell_range = f'{col_letter_target}2:{col_letter_target}{max_data_rows}'
        dv.add(cell_range)
        ws_target.add_data_validation(dv)

    # Статус (C=3) → Справочник!A
    add_ref_validation(ws, 'C', 'A', len(VALID_STATUSES_LIST), 3)
    # Тип заявки (D=4) → Справочник!C
    add_ref_validation(ws, 'D', 'C', len(app_type_names), 4)
    # ЖК (H=8) → Справочник!E
    add_ref_validation(ws, 'H', 'E', len(complexes), 8)
    # Дом (I=9) → Справочник!F
    add_ref_validation(ws, 'I', 'F', len(houses), 9)
    # Ответственный (L=12) → Справочник!B
    add_ref_validation(ws, 'L', 'B', len(rp_names), 12)
    # Источник (O=15) → Справочник!D
    add_ref_validation(ws, 'O', 'D', len(sources_list), 15)

    # --- Лист «Инструкция» ---
    ws_instr.column_dimensions['A'].width = 30
    ws_instr.column_dimensions['B'].width = 70

    instr_rows = [
        ('Инструкция по импорту заявок', ''),
        ('', ''),
        ('Режим работы', 'Описание'),
        ('Пустой «ID Заявки»', 'Создание новой заявки'),
        ('Заполненный «ID Заявки»', 'Обновление существующей заявки (изменяются только заполненные поля)'),
        ('', ''),
        ('Обязательные поля (новая заявка)', 'Описание'),
        ('ФИО Клиента', 'ФИО клиента (контакт)'),
        ('Телефон клиента', 'Номер телефона'),
        ('Тип заявки', 'Выбрать из выпадающего списка'),
        ('Ответственный', 'Выбрать из выпадающего списка'),
        ('Комментарий к заявке', 'Описание проблемы / заявки'),
        ('', ''),
        ('Обновляемые поля', 'Описание'),
        ('Статус', 'В работе / Выполнено / Частично выполнено / Закрыто / Отклонено'),
        ('Ответственный', 'ФИО ответственного'),
        ('Источник', 'Звонок / Email / Личный визит / Сайт / Другое'),
        ('Срок выполнения', 'Дата в формате ДД.ММ.ГГГГ или ГГГГ-ММ-ДД'),
        ('Последний комментарий', 'Добавляет запись в журнал заявки'),
        ('Комментарий к заявке', 'Обновляет основной комментарий заявки'),
        ('', ''),
        ('Цветовая маркировка заголовков', ''),
        ('Белый фон', 'Колонка влияет на импорт (редактируемая)'),
        ('Серый фон', 'Колонка только для чтения (игнорируется при импорте)'),
        ('', ''),
        ('Примечания', ''),
        ('', 'Можно использовать файл экспорта напрямую как файл для импорта'),
        ('', 'Колонки с серым фоном (Дата создания, № Договора, Подъезд и т.д.) не влияют на результат импорта'),
        ('', 'Для выбора значения нажмите на ячейку — появится выпадающий список'),
        ('', 'Если статус не указан при создании, устанавливается «В работе»'),
        ('', 'Если источник не указан при создании, устанавливается «Импорт»'),
    ]

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    for row_idx, (col_a, col_b) in enumerate(instr_rows, start=1):
        cell_a = ws_instr.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws_instr.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = title_font
            ws_instr.merge_cells('A1:B1')
        elif col_a and col_b == 'Описание':
            cell_a.font = section_font
            cell_b.font = section_font
        elif col_a and col_a not in ('', 'Примечания', 'Цветовая маркировка заголовков'):
            cell_a.font = Font(bold=True)
        elif col_a in ('Примечания', 'Цветовая маркировка заголовков'):
            cell_a.font = section_font
    # Цветные ячейки-примеры
    white_example_row = [r for r in range(1, len(instr_rows) + 1) if instr_rows[r-1][0] == 'Белый фон']
    gray_example_row = [r for r in range(1, len(instr_rows) + 1) if instr_rows[r-1][0] == 'Серый фон']
    if white_example_row:
        ws_instr.cell(row=white_example_row[0], column=1).fill = white_fill
    if gray_example_row:
        ws_instr.cell(row=gray_example_row[0], column=1).fill = gray_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    return send_file(buffer, as_attachment=True,
                     download_name='Шаблон_импорта_заявок.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _normalize_str(val):
    """Нормализует строку из Excel: убирает _x000D_, \r, лишние пробелы.
    openpyxl при чтении файлов, сохранённых из Excel (Windows),
    может оставлять escape-последовательности _x000D_ (\r) в тексте."""
    if val is None:
        return ''
    s = str(val)
    # openpyxl XML escape для CR
    s = s.replace('_x000D_', '')
    # Windows CRLF → LF
    s = s.replace('\r\n', '\n')
    # Оставшиеся одиночные CR
    s = s.replace('\r', '')
    return s.strip()


def _parse_import_file(filepath):
    """Парсит загруженный Excel-файл и возвращает список изменений + ошибки."""
    from .auth_utils import has_permission, get_current_user_id

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Ищем строку заголовков (содержащую 'ID Заявки')
    header_row = None
    headers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        for cell in row:
            if cell.value and str(cell.value).strip() == 'ID Заявки':
                header_row = row_idx
                break
        if header_row:
            for cell in row:
                if cell.value:
                    headers[str(cell.value).strip()] = cell.column - 1  # 0-indexed
            break

    if header_row is None:
        wb.close()
        return [], [{'row': 0, 'error': 'Не найден заголовок "ID Заявки". Убедитесь, что файл соответствует формату экспорта.'}]

    id_col = headers.get('ID Заявки')
    if id_col is None:
        wb.close()
        return [], [{'row': 0, 'error': 'Колонка "ID Заявки" не найдена в заголовках.'}]

    # Определяем индексы обновляемых колонок
    col_map = {}  # internal_name → column_index
    for excel_name, internal_name in IMPORT_UPDATABLE.items():
        if excel_name in headers:
            col_map[internal_name] = headers[excel_name]

    # Загрузим справочники
    responsible_persons = {rp.full_name.strip().lower(): rp for rp in ResponsiblePerson.query.all()}
    responsible_by_id = {rp.id: rp for rp in ResponsiblePerson.query.all()}

    # Проверяем права
    has_admin = has_permission('client-service.applications.view.all')
    current_gw_id = get_current_user_id()
    current_responsible = None
    if current_gw_id:
        current_responsible = ResponsiblePerson.query.filter_by(gateway_user_id=current_gw_id).first()

    changes = []      # обновления существующих заявок
    new_apps = []      # создание новых заявок
    errors = []
    seen_ids = set()

    # Колонки для создания новых заявок
    name_col = headers.get('ФИО Клиента')
    phone_col = headers.get('Телефон клиента')
    type_col = headers.get('Тип заявки')
    hc_col = headers.get('ЖК')
    house_col = headers.get('Дом')

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row:
            continue
        # Проверяем, есть ли хоть одно непустое значение в строке
        if all(c is None or str(c).strip() == '' for c in row):
            continue

        raw_id = row[id_col] if len(row) > id_col else None
        is_new = raw_id is None or str(raw_id).strip() == ''

        if is_new:
            # ===== СОЗДАНИЕ НОВОЙ ЗАЯВКИ =====
            def _get(col_idx):
                if col_idx is not None and len(row) > col_idx and row[col_idx] is not None:
                    return _normalize_str(row[col_idx])
                return ''

            contact_name = _get(name_col)
            contact_phone = _get(phone_col)
            app_type_name = _get(type_col)
            new_status = _get(col_map.get('status', -1)) if 'status' in col_map else ''
            responsible_name = _get(col_map.get('responsible', -1)) if 'responsible' in col_map else ''
            source_val = _get(col_map.get('source', -1)) if 'source' in col_map else ''
            comment_val = _get(col_map.get('comment', -1)) if 'comment' in col_map else ''
            hc_val = _get(hc_col)
            house_val = _get(house_col)

            # Обязательные поля
            missing = []
            if not contact_name: missing.append('ФИО Клиента')
            if not contact_phone: missing.append('Телефон клиента')
            if not app_type_name: missing.append('Тип заявки')
            if not responsible_name: missing.append('Ответственный')
            if not comment_val: missing.append('Комментарий к заявке')
            if missing:
                errors.append({'row': row_idx, 'error': f'Новая заявка: не заполнены обязательные поля: {", ".join(missing)}'})
                continue

            # Валидация статуса
            if new_status and new_status not in VALID_STATUSES:
                errors.append({'row': row_idx, 'error': f'Новая заявка: невалидный статус "{new_status}"'})
                continue

            # Поиск ответственного
            matched_rp = responsible_persons.get(responsible_name.lower())
            if not matched_rp:
                for name, rp in responsible_persons.items():
                    if responsible_name.lower() in name or name in responsible_name.lower():
                        matched_rp = rp
                        break
            if not matched_rp:
                errors.append({'row': row_idx, 'error': f'Новая заявка: ответственный "{responsible_name}" не найден'})
                continue

            # Срок выполнения
            due_date_val = None
            if 'due_date' in col_map:
                raw_due = row[col_map['due_date']] if len(row) > col_map['due_date'] else None
                if raw_due and str(raw_due).strip() not in ('', 'N/A', 'None'):
                    if isinstance(raw_due, datetime):
                        due_date_val = raw_due.isoformat()
                    else:
                        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                            try:
                                due_date_val = datetime.strptime(str(raw_due).strip(), fmt).isoformat()
                                break
                            except ValueError:
                                pass

            new_apps.append({
                'row': row_idx,
                'contact_name': contact_name,
                'contact_phone': contact_phone,
                'application_type': app_type_name,
                'status': new_status or 'В работе',
                'responsible_name': matched_rp.full_name,
                'responsible_id': matched_rp.id,
                'source': source_val or 'Импорт',
                'comment': comment_val,
                'housing_complex': hc_val,
                'house_number': house_val,
                'due_date': due_date_val,
            })
            continue

        # ===== ОБНОВЛЕНИЕ СУЩЕСТВУЮЩЕЙ ЗАЯВКИ =====
        try:
            app_id = int(raw_id)
        except (ValueError, TypeError):
            errors.append({'row': row_idx, 'error': f'Невалидный ID заявки: "{raw_id}"'})
            continue

        if app_id in seen_ids:
            errors.append({'row': row_idx, 'error': f'Дубликат ID #{app_id} в файле'})
            continue
        seen_ids.add(app_id)

        app = Application.query.get(app_id)
        if not app:
            errors.append({'row': row_idx, 'error': f'Заявка #{app_id} не найдена в базе'})
            continue

        # Проверка прав на эту заявку
        if not has_admin:
            is_responsible = current_responsible and current_responsible.id == app.responsible_person_id
            is_nc = (app.agreement_number and
                     (app.agreement_number.startswith('NC-') or app.agreement_number == 'SYSTEM-001'))
            if not (is_responsible or is_nc):
                errors.append({'row': row_idx, 'error': f'Заявка #{app_id}: нет прав на изменение'})
                continue

        row_changes = []

        # --- Статус ---
        if 'status' in col_map:
            new_val = _normalize_str(row[col_map['status']])
            if new_val and new_val != _normalize_str(app.status):
                if new_val not in VALID_STATUSES:
                    errors.append({'row': row_idx, 'error': f'Заявка #{app_id}: невалидный статус "{new_val}"'})
                    continue
                row_changes.append({
                    'field': 'Статус', 'old': app.status or '', 'new': new_val,
                    'db_field': 'status'
                })

        # --- Ответственный ---
        if 'responsible' in col_map:
            new_val = _normalize_str(row[col_map['responsible']])
            current_name = _normalize_str(app.responsible_person.full_name) if app.responsible_person else 'Не назначен'
            if new_val and new_val != current_name:
                # Матчинг по имени (case-insensitive)
                matched = responsible_persons.get(new_val.lower())
                if not matched:
                    # Частичный поиск
                    for name, rp in responsible_persons.items():
                        if new_val.lower() in name or name in new_val.lower():
                            matched = rp
                            break
                if not matched and new_val.lower() != 'не назначен':
                    errors.append({'row': row_idx, 'error': f'Заявка #{app_id}: ответственный "{new_val}" не найден'})
                    continue
                new_rp_id = matched.id if matched else None
                if new_rp_id != app.responsible_person_id:
                    row_changes.append({
                        'field': 'Ответственный', 'old': current_name,
                        'new': matched.full_name if matched else 'Не назначен',
                        'db_field': 'responsible_person_id', 'db_value': new_rp_id
                    })

        # --- Источник ---
        if 'source' in col_map:
            new_val = _normalize_str(row[col_map['source']])
            current_val = _normalize_str(app.source) or 'Не указан'
            if new_val and new_val != current_val and new_val != 'Не указан':
                row_changes.append({
                    'field': 'Источник', 'old': current_val, 'new': new_val,
                    'db_field': 'source'
                })

        # --- Срок выполнения ---
        if 'due_date' in col_map:
            raw_due = row[col_map['due_date']]
            new_due = None
            if raw_due and str(raw_due).strip() not in ('', 'N/A', 'None'):
                if isinstance(raw_due, datetime):
                    new_due = raw_due
                else:
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                        try:
                            new_due = datetime.strptime(str(raw_due).strip(), fmt)
                            break
                        except ValueError:
                            pass
                    if new_due is None:
                        errors.append({'row': row_idx, 'error': f'Заявка #{app_id}: невалидная дата дедлайна "{raw_due}"'})
                        continue

            current_due = app.due_date
            # Сравниваем только дату (без времени), т.к. Excel всегда возвращает 00:00:00
            new_due_date = new_due.date() if isinstance(new_due, datetime) else new_due
            current_due_date = current_due.date() if isinstance(current_due, datetime) else current_due
            if new_due and (not current_due or new_due_date != current_due_date):
                row_changes.append({
                    'field': 'Срок выполнения',
                    'old': current_due.strftime('%Y-%m-%d') if current_due else 'N/A',
                    'new': new_due.strftime('%Y-%m-%d'),
                    'db_field': 'due_date', 'db_value': new_due.isoformat()
                })

        # --- Комментарий к заявке ---
        if 'comment' in col_map:
            new_val = _normalize_str(row[col_map['comment']])
            current_val = _normalize_str(app.comment)
            if new_val and new_val != current_val:
                row_changes.append({
                    'field': 'Комментарий к заявке',
                    'old': current_val[:80] + ('...' if len(current_val) > 80 else ''),
                    'new': new_val[:80] + ('...' if len(new_val) > 80 else ''),
                    'db_field': 'comment', 'db_value': new_val
                })

        # --- Последний комментарий (→ ApplicationLog) ---
        if 'log_comment' in col_map:
            new_val = _normalize_str(row[col_map['log_comment']])
            last_log = ApplicationLog.query.filter_by(application_id=app_id) \
                .order_by(ApplicationLog.timestamp.desc()).first()
            current_val = _normalize_str(last_log.comment) if last_log else ''
            if new_val and new_val != current_val and new_val != 'N/A':
                row_changes.append({
                    'field': 'Новый комментарий (лог)',
                    'old': (current_val[:80] + '...') if current_val and len(current_val) > 80 else (current_val or '—'),
                    'new': new_val[:80] + ('...' if len(new_val) > 80 else ''),
                    'db_field': 'log_comment', 'db_value': new_val
                })

        if row_changes:
            changes.append({'app_id': app_id, 'row': row_idx, 'fields': row_changes})

    wb.close()
    return changes, new_apps, errors


@main.route('/applications/import', methods=['POST'])
@auth_required(permission='client-service.applications.import')
def import_applications_preview():
    """Шаг 1: загрузка Excel-файла и генерация превью изменений."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400

    file = request.files['file']
    if not file.filename or not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Допустимый формат: .xlsx'}), 400

    # Проверяем размер
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_IMPORT_FILE_SIZE:
        return jsonify({'success': False, 'error': f'Файл слишком большой ({size // (1024*1024)} МБ). Максимум 20 МБ.'}), 400

    # Сохраняем во временный файл
    import_id = str(uuid.uuid4())
    tmp_dir = os.path.join(tempfile.gettempdir(), 'crm_imports')
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, f'{import_id}.xlsx')
    file.save(tmp_path)

    try:
        changes, new_apps, errors = _parse_import_file(tmp_path)
    except Exception as e:
        os.remove(tmp_path)
        return jsonify({'success': False, 'error': f'Ошибка чтения файла: {str(e)}'}), 400

    # Считаем сводку
    total_updates = len(changes)
    total_fields = sum(len(c['fields']) for c in changes)
    total_new = len(new_apps)

    return jsonify({
        'success': True,
        'import_id': import_id,
        'summary': {
            'total_applications': total_updates,
            'total_changes': total_fields,
            'total_new': total_new,
            'total_errors': len(errors)
        },
        'changes': changes,
        'new_apps': new_apps,
        'errors': errors
    })


@main.route('/applications/import/confirm', methods=['POST'])
@auth_required(permission='client-service.applications.import')
def import_applications_confirm():
    """Шаг 2: применение изменений из загруженного файла."""
    from .auth_utils import get_or_create_local_user

    data = request.get_json()
    if not data or 'import_id' not in data:
        return jsonify({'success': False, 'error': 'import_id не указан'}), 400

    import_id = data['import_id']
    tmp_dir = os.path.join(tempfile.gettempdir(), 'crm_imports')
    tmp_path = os.path.join(tmp_dir, f'{import_id}.xlsx')

    if not os.path.exists(tmp_path):
        return jsonify({'success': False, 'error': 'Файл импорта не найден или истёк. Загрузите заново.'}), 404

    try:
        changes, new_apps, errors = _parse_import_file(tmp_path)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ошибка повторного чтения файла: {str(e)}'}), 400

    if not changes and not new_apps:
        os.remove(tmp_path)
        return jsonify({'success': True, 'applied': 0, 'message': 'Нет изменений для применения.'})

    local_user = get_or_create_local_user()
    author_id = local_user.id if local_user else None

    applied_count = 0
    applied_apps = 0
    created_count = 0

    try:
        for change_group in changes:
            app_id = change_group['app_id']
            app = Application.query.get(app_id)
            if not app:
                continue

            log_parts = []
            new_log_comment = None

            for field_change in change_group['fields']:
                db_field = field_change['db_field']
                new_value = field_change.get('db_value', field_change['new'])

                if db_field == 'status':
                    old_status = app.status
                    app.status = new_value
                    app.last_status_change = datetime.now()
                    if new_value in ('Выполнено', 'Закрыто', 'Отклонено'):
                        if not app.completed_at:
                            app.completed_at = datetime.now()
                    elif old_status in ('Выполнено', 'Закрыто', 'Отклонено'):
                        app.completed_at = None
                    log_parts.append(f'Статус: {old_status} → {new_value}')

                elif db_field == 'responsible_person_id':
                    old_name = app.responsible_person.full_name if app.responsible_person else 'Не назначен'
                    app.responsible_person_id = new_value
                    log_parts.append(f'Ответственный: {old_name} → {field_change["new"]}')

                elif db_field == 'source':
                    old_val = app.source or 'Не указан'
                    app.source = new_value
                    log_parts.append(f'Источник: {old_val} → {new_value}')

                elif db_field == 'due_date':
                    old_due = app.due_date.strftime('%Y-%m-%d') if app.due_date else 'N/A'
                    app.due_date = datetime.fromisoformat(new_value)
                    log_parts.append(f'Срок: {old_due} → {app.due_date.strftime("%Y-%m-%d")}')

                elif db_field == 'comment':
                    app.comment = new_value
                    log_parts.append('Комментарий к заявке обновлён')

                elif db_field == 'log_comment':
                    new_log_comment = new_value

                applied_count += 1

            # Создаём запись в логе
            if log_parts or new_log_comment:
                action = 'Массовый импорт: ' + '; '.join(log_parts) if log_parts else 'Комментарий через импорт'
                comment = new_log_comment or '; '.join(log_parts)
                log_entry = ApplicationLog(
                    application_id=app_id,
                    action=action,
                    comment=comment,
                    author_id=author_id
                )
                db.session.add(log_entry)

            applied_apps += 1

        # ===== СОЗДАНИЕ НОВЫХ ЗАЯВОК =====
        for new_app_data in new_apps:
            try:
                new_client, agreement_number = _create_nc_contact_and_deal(
                    new_app_data['contact_name'],
                    new_app_data['contact_phone']
                )

                # Срок выполнения: из файла или из типа заявки
                due_date = None
                if new_app_data.get('due_date'):
                    due_date = datetime.fromisoformat(new_app_data['due_date'])
                else:
                    app_type = ApplicationType.query.filter_by(name=new_app_data['application_type']).first()
                    if app_type and app_type.execution_days:
                        from datetime import timedelta
                        due_date = datetime.now() + timedelta(days=app_type.execution_days)

                new_app = Application(
                    client_id=new_client.id,
                    agreement_number=agreement_number,
                    application_type=new_app_data['application_type'],
                    comment=new_app_data['comment'],
                    status=new_app_data['status'],
                    responsible_person_id=new_app_data['responsible_id'],
                    creator_id=author_id,
                    due_date=due_date,
                    source=new_app_data['source'],
                    housing_complex=new_app_data['housing_complex'] or None,
                    house_number=new_app_data['house_number'] or None,
                )
                db.session.add(new_app)
                db.session.flush()

                db.session.add(ApplicationLog(
                    application=new_app,
                    action='Заявка создана (импорт)',
                    comment=f"Клиент: {new_app_data['contact_name']}. Ответственный: {new_app_data['responsible_name']}",
                    author_id=author_id
                ))
                created_count += 1
            except Exception as e:
                errors.append({'row': new_app_data['row'], 'error': f'Ошибка создания: {str(e)}'})

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Ошибка применения: {str(e)}'}), 500
    finally:
        # Удаляем временный файл
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return jsonify({
        'success': True,
        'applied_apps': applied_apps,
        'applied_changes': applied_count,
        'created_apps': created_count,
        'message': f'Обновлено {applied_apps} заявок ({applied_count} полей). Создано {created_count} новых заявок.'
    })


@main.route('/applications')
@auth_required(any_of=['client-service.applications.view.all', 'client-service.applications.view.own', 'client-service.applications.view.responsible'])
def applications():
    page = request.args.get('page', 1, type=int)

    # Базовый запрос с предзагрузкой связанных данных для оптимизации
    query = Application.query.options(
        joinedload(Application.client),
        joinedload(Application.responsible_person),
        joinedload(Application.creator)
    )

    # Фильтрация заявок в зависимости от разрешений
    from .auth_utils import has_permission, get_current_user_id, get_or_create_local_user
    
    # Если нет разрешения на просмотр ВСЕХ заявок, фильтруем по другим разрешениям
    if not has_permission('client-service.applications.view.all'):
        current_gateway_user_id = get_current_user_id()
        local_user = get_or_create_local_user(commit=True) if current_gateway_user_id else None
        
        if local_user:
            filter_conditions = []
            
            if has_permission('client-service.applications.view.own'):
                filter_conditions.append(Application.creator_id == local_user.id)
            
            if has_permission('client-service.applications.view.responsible'):
                responsible_person = ResponsiblePerson.query.filter_by(gateway_user_id=current_gateway_user_id).first()
                if responsible_person:
                    filter_conditions.append(Application.responsible_person_id == responsible_person.id)
            
            if filter_conditions:
                query = query.filter(or_(*filter_conditions))
            else:
                query = query.filter(Application.id == -1)
        else:
            query = query.filter(Application.id == -1)
    
    # Применяем все фильтры через общую функцию
    query, filters = _apply_app_filters(query, request.args)

    # Пагинация
    apps_paginated = query.paginate(page=page, per_page=20)
    
    # Получаем уникальные типы заявок для фильтра
    application_types = db.session.query(Application.application_type).distinct().order_by(Application.application_type).all()
    application_types = [t[0] for t in application_types if t[0]]

    # Данные для фильтров
    responsible_persons = ResponsiblePerson.query.order_by(ResponsiblePerson.full_name).all()
    app_sources = ['Звонок', 'Email', 'Личный визит', 'Сайт', 'Другое']
    app_statuses = ['В работе', 'Выполнено', 'Частично выполнено', 'Закрыто', 'Отклонено']

    # Собираем активные фильтры для пагинации
    filter_keys = ['status', 'type', 'date_from', 'date_to', 'search', 'overdue',
                   'sort', 'responsible_id', 'source', 'housing_complex', 'house_number', 'app_id']
    filter_params = {k: request.args.get(k, '') for k in filter_keys if request.args.get(k, '').strip()}

    return render_template('applications.html', 
                         applications=apps_paginated,
                         application_types=application_types,
                         responsible_persons=responsible_persons,
                         app_sources=app_sources,
                         app_statuses=app_statuses,
                         filter_params=filter_params)

@main.route('/client/<signed_int:client_id>/application/create', methods=['POST'])
@auth_required(permission='client-service.applications.create')
def create_application(client_id):
    from .auth_utils import get_or_create_local_user
    
    local_user = get_or_create_local_user()
    creator_local_id = local_user.id if local_user else None
    
    contact = EstateDealsContacts.query.get_or_404(client_id)
    form_data = request.form
    application_type_name = form_data.get('application_type')
    agreement_number, comment, responsible_person_id = form_data.get(
        'agreement_number'), form_data.get('comment'), form_data.get(
        'responsible_person_id')
    
    # Получаем источник заявки
    source = form_data.get('source', 'Звонок')
    if source == 'Другое':
        custom_source = form_data.get('custom_source', '').strip()
        if custom_source:
            source = custom_source

    if not all([agreement_number, application_type_name, comment, responsible_person_id]):
        flash('Все поля, включая ответственного, обязательны для заполнения.', 'danger')
        return redirect(url_for('main.client_card', client_id=client_id))

    # Получаем тип заявки для определения срока выполнения
    app_type = ApplicationType.query.filter_by(name=application_type_name).first()
    
    # Рассчитываем срок выполнения
    due_date = None
    if app_type and app_type.execution_days:
        from datetime import timedelta
        due_date = datetime.now() + timedelta(days=app_type.execution_days)

    new_app = Application(client_id=client_id, agreement_number=agreement_number,
                          application_type=application_type_name,
                          comment=comment, responsible_person_id=responsible_person_id,
                          creator_id=creator_local_id, due_date=due_date, source=source)
    db.session.add(new_app)

    defects_data = {}
    for key, value in request.form.items():
        if key.startswith('defects-'):
            parts = key.split('-')
            if len(parts) == 3:
                index, field = int(parts[1]), parts[2]
                defects_data.setdefault(index, {})[field] = value

    if application_type_name == 'Дефекты' and not defects_data:
        db.session.rollback()
        flash('Для заявок типа "Дефекты" необходимо добавить хотя бы один дефект.', 'danger')
        return redirect(url_for('main.client_card', client_id=client_id))

    for index, data in sorted(defects_data.items()):
        defect_type, description = data.get('type'), data.get('description')
        if defect_type and description:
            db.session.add(Defect(application=new_app, defect_type=defect_type, description=description))
        else:
            flash(f'В дефекте №{index + 1} не заполнены все поля. Он не будет сохранен.', 'warning')

    db.session.add(ApplicationLog(application=new_app, action="Заявка создана",
                                  comment=f"Назначен ответственный: {ResponsiblePerson.query.get(responsible_person_id).full_name}",
                                  author_id=creator_local_id))

    try:
        db.session.commit()
        app_instance = current_app._get_current_object()
        thr = Thread(target=send_email_async, args=[app_instance, new_app.id])
        thr.start()
        flash('Заявка успешно создана! Уведомление ответственному лицу отправляется.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при сохранении заявки в базу данных: {e}', 'danger')

    return redirect(url_for('main.client_card', client_id=client_id))


def get_or_create_system_client():
    """
    Получает системного клиента или создает его, если он не существует.
    Также создает системный договор, если его нет.
    """
    system_client = EstateDealsContacts.query.filter_by(
        contacts_buy_name="СИСТЕМНЫЙ КЛИЕНТ (для заявок без договора)"
    ).first()
    
    if not system_client:
        print("INFO: Системный клиент не найден. Создаем автоматически...")
        
        # Генерируем отрицательный ID, чтобы не конфликтовать с MySQL ID
        min_contact_id = db.session.query(db.func.min(EstateDealsContacts.id)).scalar() or 0
        new_client_id = min(min_contact_id, 0) - 1
        
        system_client = EstateDealsContacts(
            id=new_client_id,
            contacts_buy_name="СИСТЕМНЫЙ КЛИЕНТ (для заявок без договора)",
            contacts_buy_phones="000-00-00"
        )
        db.session.add(system_client)
        db.session.flush()
        print(f"INFO: Создан системный клиент с ID: {system_client.id}")
        
        # Проверяем и создаем системный договор
        system_deal = EstateDeals.query.filter_by(
            agreement_number="SYSTEM-001",
            contacts_buy_id=system_client.id
        ).first()
        
        if not system_deal:
            from datetime import date
            min_deal_id = db.session.query(db.func.min(EstateDeals.id)).scalar() or 0
            new_deal_id = min(min_deal_id, 0) - 1
            
            system_deal = EstateDeals(
                id=new_deal_id,
                agreement_number="SYSTEM-001",
                contacts_buy_id=system_client.id,
                deal_status_name="Системный",
                agreement_date=date.today(),
                deal_sum=0.0,
                finances_income_reserved=0.0
            )
            db.session.add(system_deal)
            print("INFO: Создан системный договор SYSTEM-001")
        
        db.session.commit()
    
    return system_client


@main.route('/application/create-general', methods=['POST'])
@auth_required(permission='client-service.applications.create')
def create_general_application():
    """Создание заявки без договора - создает нового клиента"""
    from .auth_utils import get_or_create_local_user
    
    local_user = get_or_create_local_user()
    creator_local_id = local_user.id if local_user else None

    
    form_data = request.form
    
    # Получаем данные из формы
    application_type_name = form_data.get('application_type')
    comment = form_data.get('comment')
    responsible_person_id = form_data.get('responsible_person_id')
    contact_name = form_data.get('contact_name', '').strip()
    contact_phone = form_data.get('contact_phone', '').strip()
    client_comment = form_data.get('client_comment', '').strip()
    
    # Получаем источник заявки
    source = form_data.get('source', 'Звонок')
    if source == 'Другое':
        custom_source = form_data.get('custom_source', '').strip()
        if custom_source:
            source = custom_source
    
    # Получаем данные о ЖК и номере дома
    housing_complex = form_data.get('housing_complex', '').strip()
    house_number = form_data.get('house_number', '').strip()

    # Проверяем обязательные поля
    if not all([application_type_name, comment, responsible_person_id, contact_name, contact_phone]):
        flash('Все поля (ФИО клиента, телефон, тип заявки, комментарий и ответственный) являются обязательными.', 'danger')
        return redirect(url_for('main.index'))

    # Создаем нового клиента для заявки без договора
    try:
        new_client, agreement_number = _create_nc_contact_and_deal(
            contact_name, contact_phone,
            client_comment=client_comment if client_comment else None
        )
        print(f"INFO: Создан новый клиент (без договора): ID={new_client.id}, ФИО={contact_name}")
    except Exception as e:
        db.session.rollback()
        print(f"ERROR: Не удалось создать клиента/договор: {e}")
        flash(f'Произошла ошибка при создании клиента: {e}', 'danger')
        return redirect(url_for('main.index'))

    # Получаем тип заявки для определения срока выполнения
    app_type = ApplicationType.query.filter_by(name=application_type_name).first()
    
    # Рассчитываем срок выполнения
    due_date = None
    if app_type and app_type.execution_days:
        from datetime import timedelta
        due_date = datetime.now() + timedelta(days=app_type.execution_days)

    # Создаем заявку с новым клиентом и договором
    new_app = Application(
        client_id=new_client.id, 
        agreement_number=agreement_number,
        application_type=application_type_name,
        comment=comment, 
        responsible_person_id=responsible_person_id,
        creator_id=creator_local_id, 
        due_date=due_date, 
        source=source,
        housing_complex=housing_complex if housing_complex else None,
        house_number=house_number if house_number else None
    )
    db.session.add(new_app)
    db.session.flush()  # Получаем ID заявки

    # Обработка дефектов
    defects_data = {}
    for key, value in request.form.items():
        if key.startswith('defects-'):
            parts = key.split('-')
            if len(parts) == 3:
                index, field = int(parts[1]), parts[2]
                defects_data.setdefault(index, {})[field] = value

    if app_type and app_type.has_defect_list and not defects_data:
        db.session.rollback()
        flash('Для заявок данного типа необходимо добавить хотя бы один дефект.', 'danger')
        return redirect(url_for('main.index'))

    for index, data in sorted(defects_data.items()):
        defect_type, description = data.get('type'), data.get('description')
        if defect_type and description:
            db.session.add(Defect(application=new_app, defect_type=defect_type, description=description))
        else:
            flash(f'В дефекте №{index + 1} не заполнены все поля. Он не будет сохранен.', 'warning')

    # Добавляем лог
    db.session.add(ApplicationLog(
        application=new_app, 
        action="Заявка создана (без договора)",
        comment=f"Создан новый клиент: {contact_name}. Назначен ответственный: {ResponsiblePerson.query.get(responsible_person_id).full_name}",
        author_id=creator_local_id
    ))

    try:
        db.session.commit()
        app_instance = current_app._get_current_object()
        thr = Thread(target=send_email_async, args=[app_instance, new_app.id])
        thr.start()
        flash(f'Заявка №{new_app.id} успешно создана для нового клиента "{contact_name}"! Уведомление ответственному лицу отправляется.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при сохранении заявки в базу данных: {e}', 'danger')

    return redirect(url_for('main.client_card', client_id=new_client.id))


@main.route('/client-service/client/<signed_int:client_id>/update_comment', methods=['POST'])
@auth_required(permission='client-service.admin.users')
def update_client_comment(client_id):
    """Обновление комментария клиента (только для админа)"""
    contact = EstateDealsContacts.query.get_or_404(client_id)
    new_comment = request.form.get('client_comment', '').strip()
    
    try:
        contact.client_comment = new_comment if new_comment else None
        db.session.commit()
        flash('Комментарий клиента успешно обновлён.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при обновлении комментария: {e}', 'danger')
    
    return redirect(url_for('main.client_card', client_id=client_id))


@main.route('/responsible')
@auth_required(permission='client-service.responsible.view')
def responsible_persons():
    persons = ResponsiblePerson.query.order_by(ResponsiblePerson.full_name).all()
    all_complex_names_tuples = db.session.query(EstateHouses.complex_name).filter(EstateHouses.complex_name.isnot(None),
                                                                                  EstateHouses.complex_name != '').distinct().order_by(
        EstateHouses.complex_name).all()
    all_complex_names = [name[0] for name in all_complex_names_tuples]

    all_application_types = ApplicationType.query.order_by(ApplicationType.name).all()
    
    # Получаем пользователей из Gateway вместо локальной БД
    from .auth_utils import get_gateway_users
    all_users = get_gateway_users()
    
    return render_template('responsible.html',
                           persons=persons,
                           all_complex_names=all_complex_names,
                           all_application_types=all_application_types,
                           all_users=all_users)


@main.route('/responsible', methods=['POST'])
@auth_required(permission='client-service.responsible.create')
def create_responsible_person():
    form_data = request.form
    new_person = ResponsiblePerson(
        full_name=form_data.get('full_name'),
        email=form_data.get('email'),
        gateway_user_id=form_data.get('user_id') or None  # Gateway User ID
    )

    selected_app_type_names = form_data.getlist('application_types')
    new_person.application_types = selected_app_type_names

    assigned_complex_names = form_data.getlist('assigned_complexes')
    if assigned_complex_names:
        new_person.assigned_complexes = EstateHouses.query.filter(
            EstateHouses.complex_name.in_(assigned_complex_names)).all()
    else:
        new_person.assigned_complexes = []
    db.session.add(new_person)
    try:
        db.session.commit()
        flash(f'Ответственное лицо "{new_person.full_name}" успешно создано.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при создании: {e}', 'danger')
    return redirect(url_for('main.responsible_persons'))


@main.route('/responsible/<int:person_id>/update', methods=['POST'])
@auth_required(permission='client-service.responsible.edit')
def update_responsible_person(person_id):
    person = ResponsiblePerson.query.get_or_404(person_id)
    form_data = request.form
    person.full_name = form_data.get('full_name')
    person.email = form_data.get('email')
    person.gateway_user_id = form_data.get('user_id') or None  # Gateway User ID

    selected_app_type_names = form_data.getlist('application_types')
    person.application_types = selected_app_type_names

    assigned_complex_names = form_data.getlist('assigned_complexes')
    if assigned_complex_names:
        person.assigned_complexes = EstateHouses.query.filter(
            EstateHouses.complex_name.in_(assigned_complex_names)).all()
    else:
        person.assigned_complexes = []
    try:
        db.session.commit()
        flash(f'Данные "{person.full_name}" успешно обновлены.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при обновлении: {e}', 'danger')
    return redirect(url_for('main.responsible_persons'))


@main.route('/responsible/<int:person_id>/delete', methods=['POST'])
@auth_required(permission='client-service.responsible.delete')
def delete_responsible_person(person_id):
    person = ResponsiblePerson.query.get_or_404(person_id)
    name = person.full_name
    db.session.delete(person)
    try:
        db.session.commit()
        flash(f'Ответственное лицо "{name}" было удалено.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении: {e}', 'danger')
    return redirect(url_for('main.responsible_persons'))


@main.route('/api/responsible')
@auth_required
def get_responsible_persons():
    complex_name, app_type = request.args.get('complex_name'), request.args.get('application_type')
    if not complex_name or not app_type:
        return jsonify([])
    sql_query_complex = text(
        """SELECT DISTINCT p.id,p.full_name,p.email,p.application_types_json FROM responsible_persons p JOIN responsible_assignments ra ON p.id=ra.responsible_person_id JOIN estate_houses eh ON eh.house_id=ra.house_id WHERE eh.complex_name LIKE :complex_name""")
    params_complex = {'complex_name': complex_name}
    result = db.session.execute(sql_query_complex, params_complex)
    persons_for_complex = result.mappings().all()
    final_persons = []
    if persons_for_complex:
        for person in persons_for_complex:
            try:
                person_app_types = json.loads(person['application_types_json'])
                if app_type in person_app_types:
                    final_persons.append(person)
            except json.JSONDecodeError:
                continue
    return jsonify([{'id': p['id'], 'name': p['full_name'], 'email': p['email']} for p in final_persons])


@main.route('/application/<int:app_id>/update_status', methods=['POST'])
@auth_required(permission='client-service.applications.status.change')
def update_application_status(app_id):
    app = Application.query.get_or_404(app_id)
    
    from .auth_utils import has_permission, get_current_user_id, get_or_create_local_user
    
    # Получаем/создаём локального пользователя для author_id
    local_user = get_or_create_local_user()
    
    # Проверяем права доступа
    has_admin_rights = has_permission('client-service.applications.view.all')
    
    # Проверяем, является ли пользователь ответственным по этой заявке
    is_responsible = False
    current_gateway_user_id = get_current_user_id()
    if current_gateway_user_id:
        # Находим ResponsiblePerson по gateway_user_id
        responsible_person = ResponsiblePerson.query.filter_by(gateway_user_id=current_gateway_user_id).first()
        if responsible_person:
            is_responsible = responsible_person.id == app.responsible_person_id

    # Для заявок без договора (NC-xxx или старый SYSTEM-001) разрешаем всем пользователям с правами доступа
    is_no_contract = (app.agreement_number and 
                      (app.agreement_number.startswith('NC-') or 
                       app.agreement_number == 'SYSTEM-001' or
                       app.client.contacts_buy_name == "СИСТЕМНЫЙ КЛИЕНТ (для заявок без договора)"))
    
    if not (has_admin_rights or is_responsible or is_no_contract):
        abort(403)  # Forbidden

    new_status, comment = request.form.get('status'), request.form.get('comment')
    # Определяем страницу возврата (карточка заявки или карточка клиента)
    redirect_to = request.form.get('next') or request.args.get('next') or url_for('main.client_card', client_id=app.client_id)
    
    if not new_status or not comment:
        flash('Для смены статуса необходимо выбрать новый статус и оставить комментарий.', 'danger')
        return redirect(redirect_to)
    if app.status == new_status:
        flash('Новый статус совпадает с текущим. Изменений не внесено.', 'info')
        return redirect(redirect_to)
    old_status = app.status
    app.status = new_status
    # Обновляем временной штамп последнего изменения статуса
    app.last_status_change = datetime.now()
    
    # Устанавливаем дату завершения, если статус финальный
    if new_status in ['Выполнено', 'Закрыто', 'Отклонено']:
        if not app.completed_at:
            app.completed_at = datetime.now()
    elif old_status in ['Выполнено', 'Закрыто', 'Отклонено'] and new_status not in ['Выполнено', 'Закрыто', 'Отклонено']:
        # Если возвращаем в работу, убираем дату завершения
        app.completed_at = None
    
    log_entry = ApplicationLog(application=app, action=f"Статус изменен: {old_status} -> {new_status}", 
                               comment=comment, author_id=local_user.id if local_user else None)
    db.session.add(log_entry)
    try:
        db.session.commit()
        flash(f'Статус заявки #{app.id} успешно изменен на "{new_status}".', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при изменении статуса: {e}', 'danger')
    return redirect(redirect_to)


@main.route('/api/application/<int:app_id>/logs')
@auth_required
def get_application_logs(app_id):
    Application.query.get_or_404(app_id)
    logs = ApplicationLog.query.filter_by(application_id=app_id).options(
        joinedload(ApplicationLog.author)
    ).order_by(ApplicationLog.timestamp.desc()).all()
    return jsonify(
        [{'timestamp': log.timestamp.strftime('%d.%m.%Y %H:%M:%S'), 
          'action': log.action, 
          'comment': log.comment,
          'author': log.author.username if log.author else 'Система'} for log in logs])


@main.route('/reports')
@auth_required(permission='client-service.reports.view')
def reports():
    return render_template('reports.html')


@main.route('/reports/download', methods=['POST'])
@auth_required(permission='client-service.reports.download')
def download_report():
    start_date_str, end_date_str = request.form.get('start_date'), request.form.get('end_date')
    if not start_date_str or not end_date_str:
        flash('Необходимо указать и начальную, и конечную дату.', 'danger')
        return redirect(url_for('main.reports'))
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        flash('Неверный формат даты.', 'danger')
        return redirect(url_for('main.reports'))

    apps = Application.query.options(joinedload(Application.client), joinedload(Application.responsible_person),
                                     joinedload(Application.defects)).filter(Application.created_at >= start_date,
                                                                             Application.created_at <= end_date).order_by(
        Application.created_at.desc()).all()
    if not apps:
        flash('За указанный период не найдено ни одной заявки.', 'info')
        return redirect(url_for('main.reports'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по заявкам"
    headers = ["ID Заявки", "Дата создания", "Статус", "Тип заявки", "№ Договора", "ФИО Клиента", "Телефон клиента",
               "ЖК", "Дом", "Подъезд", "Номер квартиры",
               "Ответственный", "Email ответственного", "Источник", "Срок выполнения", "Дата завершения", "Просрочено",
               "Дата последнего изменения", "Последний комментарий", "Комментарий к заявке", "Дефекты (Тип: Комментарий)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for app in apps:
        # Получаем данные о недвижимости через связи
        complex_name = "N/A"
        house_name = "N/A"
        entrance = "N/A"
        flat_num = "N/A"
        
        # Ищем сделку по номеру договора
        if app.client and app.agreement_number:
            deal = EstateDeals.query.filter_by(
                agreement_number=app.agreement_number,
                contacts_buy_id=app.client.id
            ).first()
            
            if deal and deal.sell:
                sell = deal.sell
                entrance = sell.geo_house_entrance if sell.geo_house_entrance else "N/A"
                flat_num = sell.geo_flatnum if sell.geo_flatnum else "N/A"
                
                if sell.house:
                    complex_name = sell.house.complex_name if sell.house.complex_name else "N/A"
                    house_name = sell.house.name if sell.house.name else "N/A"
        
        defects_str = "; ".join([f"{d.defect_type}: {d.description}" for d in app.defects])
        
        # Проверяем просроченность
        is_overdue = "Да" if app.is_overdue else "Нет"
        
        # Получаем последний комментарий из логов
        last_log = ApplicationLog.query.filter_by(application_id=app.id).order_by(ApplicationLog.timestamp.desc()).first()
        last_comment = last_log.comment if last_log else "N/A"
        
        row_data = [app.id, app.created_at.strftime('%Y-%m-%d %H:%M:%S'), app.status, app.application_type,
                    app.agreement_number, app.client.contacts_buy_name if app.client else "N/A",
                    app.client.contacts_buy_phones if app.client else "N/A",
                    complex_name, house_name, entrance, flat_num,
                    app.responsible_person.full_name if app.responsible_person else "Не назначен",
                    app.responsible_person.email if app.responsible_person else "N/A",
                    app.source if app.source else "Не указан",
                    app.due_date.strftime('%Y-%m-%d') if app.due_date else "N/A",
                    app.completed_at.strftime('%Y-%m-%d %H:%M:%S') if app.completed_at else "N/A",
                    is_overdue,
                    app.last_status_change.strftime('%Y-%m-%d %H:%M:%S') if app.last_status_change else "N/A",
                    last_comment,
                    app.comment, defects_str]
        ws.append(row_data)

    for col in ws.columns:
        max_length = 0
        # Получаем букву колонки из строки заголовков (3-я строка, индекс 2),
        # чтобы избежать ошибки с объединенной ячейкой в первой строке.
        column = col[2].column_letter
        for cell in col:
            try:
                if cell.value:  # Проверяем, что ячейка не пустая
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"report_{start_date_str}_to_{end_date_str}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main.route('/reports/download-completed', methods=['POST'])
@auth_required(permission='client-service.reports.download')
def download_completed_report():
    """Генерирует отчет по завершенным заявкам за указанный период"""
    start_date_str, end_date_str = request.form.get('start_date'), request.form.get('end_date')
    if not start_date_str or not end_date_str:
        flash('Необходимо указать и начальную, и конечную дату.', 'danger')
        return redirect(url_for('main.reports'))
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        flash('Неверный формат даты.', 'danger')
        return redirect(url_for('main.reports'))

    # Фильтруем только завершенные заявки по дате завершения
    apps = Application.query.options(
        joinedload(Application.client), 
        joinedload(Application.responsible_person),
        joinedload(Application.defects),
        joinedload(Application.creator)
    ).filter(
        Application.completed_at >= start_date,
        Application.completed_at <= end_date,
        Application.status.in_(['Выполнено', 'Закрыто', 'Отклонено'])
    ).order_by(Application.completed_at.desc()).all()
    
    if not apps:
        flash('За указанный период не найдено ни одной завершенной заявки.', 'info')
        return redirect(url_for('main.reports'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Завершенные заявки"
    headers = ["ID Заявки", "Дата создания", "Дата завершения", "Время выполнения (дней)", 
               "Статус", "Тип заявки", "№ Договора", "ФИО Клиента", "Телефон клиента",
               "ЖК", "Дом", "Подъезд", "Номер квартиры",
               "Ответственный", "Email ответственного", "Создатель заявки", "Источник", 
               "Была просрочена", "Дата последнего изменения", "Последний комментарий", 
               "Комментарий к заявке", "Дефекты (Тип: Комментарий)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for app in apps:
        # Получаем данные о недвижимости через связи
        complex_name = "N/A"
        house_name = "N/A"
        entrance = "N/A"
        flat_num = "N/A"
        
        # Ищем сделку по номеру договора
        if app.client and app.agreement_number:
            deal = EstateDeals.query.filter_by(
                agreement_number=app.agreement_number,
                contacts_buy_id=app.client.id
            ).first()
            
            if deal and deal.sell:
                sell = deal.sell
                entrance = sell.geo_house_entrance if sell.geo_house_entrance else "N/A"
                flat_num = sell.geo_flatnum if sell.geo_flatnum else "N/A"
                
                if sell.house:
                    complex_name = sell.house.complex_name if sell.house.complex_name else "N/A"
                    house_name = sell.house.name if sell.house.name else "N/A"
        
        defects_str = "; ".join([f"{d.defect_type}: {d.description}" for d in app.defects])
        
        # Рассчитываем время выполнения в днях
        execution_time = "N/A"
        if app.completed_at and app.created_at:
            time_diff = app.completed_at - app.created_at
            execution_time = time_diff.days
        
        # Проверяем, была ли заявка просрочена
        was_overdue = "Нет"
        if app.due_date and app.completed_at:
            if app.completed_at > app.due_date:
                was_overdue = "Да"
        
        # Получаем последний комментарий из логов
        last_log = ApplicationLog.query.filter_by(application_id=app.id).order_by(ApplicationLog.timestamp.desc()).first()
        last_comment = last_log.comment if last_log else "N/A"
        
        row_data = [
            app.id, 
            app.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            app.completed_at.strftime('%Y-%m-%d %H:%M:%S') if app.completed_at else "N/A",
            execution_time,
            app.status, 
            app.application_type,
            app.agreement_number, 
            app.client.contacts_buy_name if app.client else "N/A",
            app.client.contacts_buy_phones if app.client else "N/A",
            complex_name, house_name, entrance, flat_num,
            app.responsible_person.full_name if app.responsible_person else "Не назначен",
            app.responsible_person.email if app.responsible_person else "N/A",
            app.creator.username if app.creator else "Система",
            app.source if app.source else "Не указан",
            was_overdue,
            app.last_status_change.strftime('%Y-%m-%d %H:%M:%S') if app.last_status_change else "N/A",
            last_comment,
            app.comment, 
            defects_str
        ]
        ws.append(row_data)

    # Автоматическая настройка ширины колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"completed_applications_{start_date_str}_to_{end_date_str}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main.route('/deadlines/upload', methods=['GET', 'POST'])
@auth_required(permission='client-service.admin.settings')
def upload_deadlines():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не был выбран.', 'danger')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('Файл не был выбран.', 'danger')
            return redirect(request.url)
        if file and file.filename.endswith('.xlsx'):
            try:
                workbook = load_workbook(file)
                sheet = workbook.active

                updated_count = 0
                not_found_count = 0
                error_count = 0

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row): continue

                    house_name = row[0]
                    house_date_raw = row[1]
                    apartments_date_raw = row[2]

                    if not house_name:
                        flash(f'Ошибка в строке {row_idx}: Название Дома не указано. Строка пропущена.', 'warning')
                        error_count += 1
                        continue

                    house = EstateHouses.query.filter_by(name=house_name).first()

                    if not house:
                        flash(f'Дом с названием "{house_name}" из строки {row_idx} не найден в базе данных.', 'warning')
                        not_found_count += 1
                        continue

                    try:
                        if house_date_raw:
                            parsed_house_date = house_date_raw.date() if isinstance(house_date_raw,
                                                                                    datetime) else datetime.strptime(
                                str(house_date_raw), '%d.%m.%Y').date()
                            house.warranty_house_end_date = parsed_house_date

                        if apartments_date_raw:
                            parsed_apartments_date = apartments_date_raw.date() if isinstance(apartments_date_raw,
                                                                                              datetime) else datetime.strptime(
                                str(apartments_date_raw), '%d.%m.%Y').date()
                            house.warranty_apartments_end_date = parsed_apartments_date

                        updated_count += 1

                    except (ValueError, TypeError):
                        flash(
                            f'Ошибка в строке {row_idx} для дома "{house_name}": Неверный формат даты. Ожидается ДД.ММ.ГГГГ. Строка пропущена.',
                            'warning')
                        error_count += 1
                        continue

                db.session.commit()

                success_message = f'Обработка файла завершена. '
                if updated_count > 0:
                    success_message += f'Обработано {updated_count} строк. Данные сохранены. '
                if not_found_count > 0:
                    success_message += f'Не найдено {not_found_count} домов. '
                if error_count > 0:
                    success_message += f'Ошибок в данных: {error_count}.'

                flash(success_message, 'success' if error_count == 0 and not_found_count == 0 else 'info')

            except Exception as e:
                db.session.rollback()
                flash(f'Произошла критическая ошибка при обработке файла: {e}', 'danger')

            return redirect(url_for('main.upload_deadlines'))

        else:
            flash('Неверный формат файла. Пожалуйста, загрузите файл .xlsx', 'danger')
            return redirect(request.url)

    return render_template('upload_deadlines.html')


@main.route('/deadlines/template')
@auth_required(permission='client-service.admin.settings')
def download_deadlines_template():
    try:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Шаблон сроков гарантии"

        headers = ["Название Дома", "Срок гарантии по Дому (формат ДД.ММ.ГГГГ)",
                   "Срок гарантии по Квартирам (формат ДД.ММ.ГГГГ)"]
        ws.append(headers)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30

        unique_house_names_tuples = db.session.query(EstateHouses.name).filter(EstateHouses.name.isnot(None),
                                                                               EstateHouses.name != '').distinct().order_by(
            EstateHouses.name).all()

        for name_tuple in unique_house_names_tuples:
            house_name = name_tuple[0]
            house = EstateHouses.query.filter_by(name=house_name).first()
            if house:
                house_date_str = house.warranty_house_end_date.strftime(
                    '%d.%m.%Y') if house.warranty_house_end_date else ''
                apartments_date_str = house.warranty_apartments_end_date.strftime(
                    '%d.%m.%Y') if house.warranty_apartments_end_date else ''
                ws.append([house.name, house_date_str, apartments_date_str])

        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name='shablon_srokov_garantii_po_domam.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Не удалось сгенерировать шаблон. Ошибка: {e}', 'danger')
        return redirect(url_for('main.upload_deadlines'))


@main.route('/admin/email-logs')
@auth_required(permission='client-service.admin.logs')
def email_logs():
    page = request.args.get('page', 1, type=int)
    logs = EmailLog.query.order_by(EmailLog.timestamp.desc()).paginate(page=page, per_page=20)
    return render_template('email_logs.html', logs=logs)


@main.route('/admin/defect-types', methods=['GET', 'POST'])
@auth_required(permission='client-service.admin.settings')
def manage_defect_types():
    if request.method == 'POST':
        new_type_name = request.form.get('name')
        if new_type_name:
            existing_type = DefectType.query.filter_by(name=new_type_name).first()
            if not existing_type:
                new_type = DefectType(name=new_type_name)
                db.session.add(new_type)
                db.session.commit()
                flash(f'Тип дефекта "{new_type_name}" успешно добавлен.', 'success')
            else:
                flash(f'Тип дефекта "{new_type_name}" уже существует.', 'warning')
        else:
            flash('Название типа дефекта не может быть пустым.', 'danger')
        return redirect(url_for('main.manage_defect_types'))

    defect_types = DefectType.query.order_by(DefectType.name).all()
    return render_template('manage_defect_types.html', defect_types=defect_types)


@main.route('/admin/defect-types/<int:type_id>/delete', methods=['POST'])
@auth_required(permission='client-service.admin.settings')
def delete_defect_type(type_id):
    type_to_delete = DefectType.query.get_or_404(type_id)
    try:
        db.session.delete(type_to_delete)
        db.session.commit()
        flash(f'Тип дефекта "{type_to_delete.name}" удален.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось удалить тип дефекта. Возможно, он где-то используется. Ошибка: {e}', 'danger')
    return redirect(url_for('main.manage_defect_types'))


@main.route('/admin/application-types', methods=['GET', 'POST'])
@auth_required(permission='client-service.admin.settings')
def manage_application_types():
    if request.method == 'POST':
        name = request.form.get('name')
        # --- ИЗМЕНЕНИЕ: Получаем значение чекбокса ---
        has_defect_list = 'has_defect_list' in request.form
        
        # Получаем срок выполнения
        execution_days = request.form.get('execution_days', type=int, default=3)
        if execution_days < 1:
            execution_days = 3  # Минимальный срок - 1 день

        if not name:
            flash('Название типа заявки не может быть пустым.', 'danger')
            return redirect(url_for('main.manage_application_types'))

        if 'template' not in request.files:
            flash('Файл шаблона не был выбран.', 'danger')
            return redirect(url_for('main.manage_application_types'))

        file = request.files['template']

        if file.filename == '':
            flash('Файл шаблона не был выбран.', 'danger')
            return redirect(url_for('main.manage_application_types'))

        if file and file.filename.endswith('.docx'):
            filename = secure_filename(file.filename)
            upload_folder = os.path.join(current_app.root_path, 'word_templates')
            os.makedirs(upload_folder, exist_ok=True)

            file.save(os.path.join(upload_folder, filename))

            # --- ИЗМЕНЕНИЕ: Сохраняем новый флаг в БД ---
            new_app_type = ApplicationType(name=name, template_filename=filename, has_defect_list=has_defect_list, execution_days=execution_days)
            db.session.add(new_app_type)
            db.session.commit()
            flash(f'Тип заявки "{name}" с шаблоном "{filename}" и сроком выполнения {execution_days} дн.(-я) успешно создан.', 'success')
        else:
            flash('Разрешены только файлы формата .docx', 'danger')

        return redirect(url_for('main.manage_application_types'))
    template_tags = [
        {'tag': '{{ fio_otvetstvenni }}', 'desc': 'ФИО ответственного сотрудника'},
        {'tag': '{{ request_id }}', 'desc': 'ID созданной заявки'},
        {'tag': '{{ today_date }}', 'desc': 'Текущая дата в формате ДД.ММ.ГГГГ'},
        {'tag': '{{ agreement_number }}', 'desc': 'Номер договора, к которому привязана заявка'},
        {'tag': '{{ client_fio }}', 'desc': 'ФИО клиента'},
        {'tag': '{{ complex_name }}', 'desc': 'Название ЖК (жилого комплекса)'},
        {'tag': '{{ house_name }}', 'desc': 'Название/номер дома'},
        {'tag': '{{ podiezd }}', 'desc': 'Номер подъезда'},
        {'tag': '{{ flat_num }}', 'desc': 'Номер квартиры'},
        {'tag': '{{ comment }}', 'desc': 'Общий комментарий к заявке'},
        {'tag': '{{ client_phone_number }}', 'desc': 'Контактный телефон клиента'},
        {'tag': '{% for defect in defects %}...{% endfor %}', 'desc': 'Цикл для перечисления дефектов'},
        {'tag': '{{ defect.defect_type }}', 'desc': 'Тип дефекта (внутри цикла)'},
        {'tag': '{{ defect.comment }}', 'desc': 'Комментарий к конкретному дефекту (внутри цикла)'},
    ]

    app_types = ApplicationType.query.order_by(ApplicationType.name).all()
    return render_template('manage_application_types.html', app_types=app_types, template_tags=template_tags)


@main.route('/admin/application-types/<int:type_id>/download')
@auth_required(permission='client-service.admin.settings')
def download_application_template(type_id):
    """Скачивание шаблона Word для типа заявки"""
    app_type = ApplicationType.query.get_or_404(type_id)
    
    if not app_type.template_filename:
        flash('У данного типа заявки нет шаблона.', 'warning')
        return redirect(url_for('main.manage_application_types'))
    
    template_path = os.path.join(current_app.root_path, 'word_templates', app_type.template_filename)
    
    if not os.path.exists(template_path):
        flash('Файл шаблона не найден на сервере.', 'danger')
        return redirect(url_for('main.manage_application_types'))
    
    # Формируем имя файла для скачивания
    download_name = f"template_{app_type.name}_{app_type.template_filename}"
    # Заменяем пробелы на подчеркивания для корректного имени файла
    download_name = download_name.replace(' ', '_')
    
    return send_file(
        template_path,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )


@main.route('/admin/application-types/<int:type_id>/delete', methods=['POST'])
@auth_required(permission='client-service.admin.settings')
def delete_application_type(type_id):
    app_type_to_delete = ApplicationType.query.get_or_404(type_id)

    if app_type_to_delete.template_filename:
        try:
            template_path = os.path.join(current_app.root_path, 'word_templates', app_type_to_delete.template_filename)
            if os.path.exists(template_path):
                os.remove(template_path)
        except OSError as e:
            flash(f'Не удалось удалить файл шаблона {app_type_to_delete.template_filename}. Ошибка: {e}', 'warning')

    db.session.delete(app_type_to_delete)
    db.session.commit()
    flash(f'Тип заявки "{app_type_to_delete.name}" удален.', 'success')
    return redirect(url_for('main.manage_application_types'))


@main.route('/application/<int:app_id>/delete', methods=['POST'])
@auth_required(permission='client-service.applications.delete')
def delete_application(app_id):
    """
    Удаляет заявку и все связанные с ней данные (дефекты, логи).
    Доступно только администраторам.
    """
    app_to_delete = Application.query.get_or_404(app_id)

    try:
        # База данных настроена с каскадным удалением (cascade="all, delete-orphan"),
        # поэтому связанные дефекты и логи удалятся автоматически.
        db.session.delete(app_to_delete)
        db.session.commit()
        flash(f'Заявка #{app_id} была успешно удалена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при удалении заявки #{app_id}: {e}', 'danger')

    # Перенаправляем пользователя обратно на страницу со списком заявок
    return redirect(url_for('main.applications'))


@main.route('/api/housing-complexes', methods=['GET'])
@auth_required()
def get_housing_complexes():
    """API для получения списка уникальных ЖК"""
    try:
        result = db.session.execute(
            db.text('SELECT DISTINCT complex_name FROM estate_houses WHERE complex_name IS NOT NULL ORDER BY complex_name')
        ).fetchall()
        complexes = [row[0] for row in result]
        return jsonify({'complexes': complexes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main.route('/api/house-numbers', methods=['GET'])
@auth_required()
def get_house_numbers():
    """API для получения списка домов для выбранного ЖК"""
    complex_name = request.args.get('complex')
    if not complex_name:
        return jsonify({'error': 'Complex name is required'}), 400
    
    try:
        result = db.session.execute(
            db.text('SELECT DISTINCT name FROM estate_houses WHERE complex_name = :complex ORDER BY name'),
            {'complex': complex_name}
        ).fetchall()
        houses = [row[0] for row in result]
        return jsonify({'houses': houses})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


